"""Utility functions to parse Excel files for product and purchase data."""

from __future__ import annotations

import io
from datetime import datetime, date
from typing import Any

from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _to_str(val: Any) -> str | None:
    """Return *val* as a stripped string, or None if empty / None."""
    if val is None:
        return None
    return str(val).strip()


def _resolve_merged_cells(ws: Any, max_col: int) -> list[list[Any]]:
    """
    Return a 2-D list of *computed* cell values with merged ranges filled
    down / right so that every logical row has all values.

    openpyxl's ``iter_rows`` only gives the top-left cell for each merged
    range; we propagate the value to every cell in the range.
    """
    # Build a mapping  (row, col) -> value  using iter_rows (which only
    # returns the top-left cell of a merged range).
    cell_map: dict[tuple[int, int], Any] = {}
    for row in ws.iter_rows(min_row=1, max_col=max_col, values_only=False):
        for cell in row:
            cell_map[(cell.row, cell.column)] = cell.value

    # Now fill every cell that sits inside a merged range.
    for merge in ws.merged_cells.ranges:
        min_row, min_col = merge.min_row, merge.min_col
        max_row, max_col_m = merge.max_row, merge.max_col
        top_left_value = cell_map.get((min_row, min_col))
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col_m + 1):
                cell_map[(r, c)] = top_left_value

    # Build the final matrix (1-indexed rows).
    max_row_num = ws.max_row or 0
    rows: list[list[Any]] = []
    for r in range(1, max_row_num + 1):
        row_vals: list[Any] = []
        for c in range(1, max_col + 1):
            row_vals.append(cell_map.get((r, c)))
        rows.append(row_vals)
    return rows


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def _detect_column_mapping(header_row: list[Any]) -> dict[str, int]:
    """Auto-detect column indices from header row text.

    Returns a dict mapping logical name -> column index (0-based).
    Falls back to fixed positions if no headers match.
    """
    # Normalize header strings
    headers = [str(h).strip().lower() if h else "" for h in header_row]

    mapping: dict[str, int] = {}

    # SKU column
    for i, h in enumerate(headers):
        if h in ("sku", "sku编码", "sku code", "新sku", "新品sku"):
            mapping["sku"] = i
            break
    # product_name column
    for i, h in enumerate(headers):
        if h in ("产品名称", "商品名称", "product_name", "品名", "名称", "中文名称"):
            mapping["product_name"] = i
            break
    # old_sku column
    for i, h in enumerate(headers):
        if h in ("旧sku", "旧品sku", "old_sku", "原sku"):
            mapping["old_sku"] = i
            break
    # status column
    for i, h in enumerate(headers):
        if h in ("状态", "status", "商品状态"):
            mapping["status"] = i
            break

    # Fallback to fixed positions if detection failed
    mapping.setdefault("sku", 1)
    mapping.setdefault("product_name", 2)
    mapping.setdefault("old_sku", 3)
    mapping.setdefault("status", 4)

    return mapping


def parse_products_excel(file_bytes: bytes) -> list[dict[str, Any]]:
    """
    Parse a product-SKU Excel file with auto-detected column mapping.

    Returns a list of dicts with keys:
        row_num, sku, product_name, old_sku, status

    Rows are **skipped** when:
        * product_name is empty / None
        * status equals "说明SKU"
    """
    wb = load_workbook(io.BytesIO(file_bytes), read_only=False, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        return []

    max_col = ws.max_column or 5
    rows = _resolve_merged_cells(ws, max_col)
    wb.close()

    if not rows:
        return []

    # Detect column mapping from header row
    col_map = _detect_column_mapping(rows[0])

    products: list[dict[str, Any]] = []
    for data_row_idx, row in enumerate(rows):
        # Row 0 is the header – skip.
        if data_row_idx == 0:
            continue
        if len(row) < 2:
            continue

        sku = (_to_str(row[col_map["sku"]]) if col_map["sku"] < len(row) else None) or ""
        product_name = _to_str(row[col_map["product_name"]]) if col_map["product_name"] < len(row) else None
        old_sku = (_to_str(row[col_map["old_sku"]]) if col_map["old_sku"] < len(row) else None) or ""
        status = _to_str(row[col_map["status"]]) if col_map["status"] < len(row) else None

        # Normalize SKU fields to uppercase
        sku = sku.strip().upper()
        old_sku = old_sku.strip().upper()

        # Skip rows without a product name.
        if not product_name:
            continue
        # Skip rows with empty SKU (likely section headers / group titles).
        if not sku:
            continue
        # Skip rows with status "说明SKU".
        if status == "说明SKU":
            continue

        row_num = data_row_idx + 1

        products.append(
            {
                "row_num": row_num,
                "sku": sku or "",
                "product_name": product_name,
                "old_sku": old_sku or "",
                "status": status or "active",
            }
        )

    return products


def _detect_purchase_column_mapping(header_row: list[Any]) -> dict[str, int]:
    """Auto-detect purchase Excel column indices from header row text.

    Supports real headers like: 销售日期 / 客户ID / SKU / 数量 / 中文名称
    Also supports generic headers: user_id / sku / product_name / quantity / purchase_date
    """
    headers = [str(h).strip().lower() if h else "" for h in header_row]
    mapping: dict[str, int] = {}

    for i, h in enumerate(headers):
        if h in ("销售日期", "purchase_date", "日期", "购买日期", "订单日期"):
            mapping["purchase_date"] = i
        elif h in ("客户id", "客户编号", "user_id", "用户id", "用户编号"):
            mapping["user_id"] = i
        elif h in ("sku", "sku编码", "sku code", "新sku"):
            mapping["sku"] = i
        elif h in ("数量", "quantity", "采购数量", "购买数量"):
            mapping["quantity"] = i
        elif h in ("中文名称", "产品名称", "商品名称", "product_name", "品名", "名称"):
            mapping["product_name"] = i
        elif h in ("原始sku", "原sku", "original_sku", "旧sku"):
            mapping["original_sku"] = i

    # Fallback to fixed positions
    mapping.setdefault("purchase_date", 0)
    mapping.setdefault("user_id", 1)
    mapping.setdefault("sku", 2)
    mapping.setdefault("quantity", 3)
    mapping.setdefault("product_name", 4)

    return mapping


def parse_purchases_excel(file_bytes: bytes) -> list[dict[str, Any]]:
    """
    Parse a customer-purchase Excel file with auto-detected column mapping.

    Supports real headers like: 销售日期 / 客户ID / SKU / 数量 / 中文名称

    Returns a list of dicts with keys:
        user_id, sku, product_name, quantity, purchase_date, original_sku
    """
    wb = load_workbook(io.BytesIO(file_bytes), read_only=False, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        return []

    max_col = ws.max_column or 6
    rows = _resolve_merged_cells(ws, max_col)
    wb.close()

    if not rows:
        return []

    col_map = _detect_purchase_column_mapping(rows[0])

    purchases: list[dict[str, Any]] = []
    last_purchase_date: date | None = None
    last_user_id: str = ""

    for data_row_idx, row in enumerate(rows):
        # Row 0 is the header – skip.
        if data_row_idx == 0:
            continue
        if len(row) < 2:
            continue

        def _get(key: str) -> Any:
            idx = col_map.get(key, -1)
            return row[idx] if 0 <= idx < len(row) else None

        raw_purchase_date = _get("purchase_date")
        user_id = (_to_str(_get("user_id")) or "").strip().upper()
        sku = (_to_str(_get("sku")) or "").strip().upper()
        product_name = _to_str(_get("product_name")) or ""
        raw_quantity = _get("quantity")
        original_sku = _to_str(_get("original_sku")) if "original_sku" in col_map else ""

        # Quantity – default to 1 if missing or non-numeric.
        try:
            quantity = int(float(raw_quantity)) if raw_quantity is not None else 1
        except (ValueError, TypeError):
            quantity = 1

        # Purchase date – may be a datetime/date object or a string.
        purchase_date_val = _parse_date(raw_purchase_date)
        if purchase_date_val is not None:
            last_purchase_date = purchase_date_val
        else:
            # Inherit from previous row (handles merged cells).
            purchase_date_val = last_purchase_date

        if purchase_date_val is None:
            purchase_date_val = date.today()

        # user_id – inherit from previous row if empty (merged cell)
        if user_id:
            last_user_id = user_id
        else:
            user_id = last_user_id

        purchases.append(
            {
                "user_id": user_id,
                "sku": sku,
                "product_name": product_name,
                "quantity": quantity,
                "purchase_date": purchase_date_val,
                "original_sku": original_sku,
            }
        )

    return purchases


# ---------------------------------------------------------------------------
# Date parsing helper
# ---------------------------------------------------------------------------

def _parse_date(value: Any) -> date | None:
    """Try to coerce *value* into a :class:`datetime.date`, or return None."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y年%m月%d日"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
        # Last-resort: try ISO format.
        try:
            return datetime.fromisoformat(value).date()
        except (ValueError, TypeError):
            return None
    # Some spreadsheets give an int (serial date). openpyxl usually converts
    # these already, but handle it just in case.
    return None
